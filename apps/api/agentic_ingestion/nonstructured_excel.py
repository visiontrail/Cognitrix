"""Workbook structural inspector.

Produces a neutral structural summary of an uploaded Excel workbook so the
ingestion agent can decide on its own whether the workbook is a flat
structured table, a human-readable matrix, or something in between, and
how to decompose it into analysis-ready tables.

This module intentionally does NOT emit business names, table names,
business types, human labels, or any pre-baked catalog seed. All such
naming and decomposition decisions are made by the LLM based on the
observed content.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


MAX_TOP_ROWS = 15
MAX_TOP_COLUMNS = 30
MAX_MERGED_RANGE_SAMPLES = 8
MAX_CELL_TEXT_LEN = 120


def inspect_workbook_structure(
    *,
    workbook_bytes: bytes | None = None,
    workbook_path: Path | None = None,
) -> dict[str, Any] | None:
    """Return a neutral structural summary of the workbook.

    Each sheet entry exposes the raw structural signals the ingestion agent
    needs to self-classify the layout (flat table vs human-readable matrix
    vs mixed). The summary deliberately carries no business naming.
    """

    if workbook_bytes is None and workbook_path is None:
        raise ValueError("workbook_bytes or workbook_path is required")

    source: BytesIO | Path
    if workbook_bytes is not None:
        source = BytesIO(workbook_bytes)
    else:
        source = Path(workbook_path or "")

    try:
        workbook = load_workbook(source, data_only=True, read_only=False)
    except Exception:
        return None

    sheets_summary: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        try:
            sheets_summary.append(_summarize_sheet(worksheet))
        except Exception:
            continue

    if not sheets_summary:
        return None

    return {"sheets": sheets_summary}


TOP_METADATA_ROW_WINDOW = 6


def _summarize_sheet(worksheet: Worksheet) -> dict[str, Any]:
    merged_lookup = _build_merged_value_lookup(worksheet)
    max_row = int(worksheet.max_row or 0)
    max_col = int(worksheet.max_column or 0)

    preview_rows = min(max_row, MAX_TOP_ROWS)
    preview_cols = min(max_col, MAX_TOP_COLUMNS)
    top_rows_preview: list[list[str]] = []
    for row in range(1, preview_rows + 1):
        row_cells = [
            _cell_text(worksheet, row, col, merged_lookup)
            for col in range(1, preview_cols + 1)
        ]
        top_rows_preview.append(row_cells)

    merged_ranges = list(worksheet.merged_cells.ranges)
    merged_range_samples = [str(rng) for rng in merged_ranges[:MAX_MERGED_RANGE_SAMPLES]]

    has_merged_cells = len(merged_ranges) > 0
    top_horizontal_merge_count = sum(
        1
        for rng in merged_ranges
        if rng.min_row <= TOP_METADATA_ROW_WINDOW and rng.max_col > rng.min_col
    )
    has_stacked_top_metadata = top_horizontal_merge_count >= 2
    likely_layout = _classify_layout(
        has_merged_cells=has_merged_cells,
        has_stacked_top_metadata=has_stacked_top_metadata,
        max_row=max_row,
    )

    return {
        "sheet_name": worksheet.title,
        "max_row": max_row,
        "max_column": max_col,
        "merged_cell_count": len(merged_ranges),
        "merged_cell_range_samples": merged_range_samples,
        "top_horizontal_merge_count": top_horizontal_merge_count,
        "top_rows_preview": top_rows_preview,
        "structural_signals": {
            "has_merged_cells": has_merged_cells,
            "has_stacked_top_metadata": has_stacked_top_metadata,
            "likely_layout": likely_layout,
        },
    }


def _build_merged_value_lookup(worksheet: Worksheet) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, col)] = top_left
    return values


def _cell_text(
    worksheet: Worksheet,
    row: int,
    col: int,
    merged_values: dict[tuple[int, int], Any],
) -> str:
    raw = worksheet.cell(row, col).value
    if raw is None:
        raw = merged_values.get((row, col))
    if raw is None:
        return ""
    text = str(raw).replace("\r\n", "\n").replace("\r", "\n")
    text = " / ".join(line.strip() for line in text.split("\n") if line.strip())
    if len(text) > MAX_CELL_TEXT_LEN:
        text = text[: MAX_CELL_TEXT_LEN - 1] + "…"
    return text


def _classify_layout(
    *,
    has_merged_cells: bool,
    has_stacked_top_metadata: bool,
    max_row: int,
) -> str:
    if max_row <= 0:
        return "empty"
    if has_stacked_top_metadata:
        return "human_readable_matrix"
    if not has_merged_cells:
        return "flat_table"
    return "mixed_or_unknown"
